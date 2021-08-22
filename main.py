import openpyxl

import requests 
from bs4 import BeautifulSoup

import fake_useragent
from random import choice

base_url = 'https://www.avito.ru'
proxies = open('work_proxies.txt').read().split('\n')
all_cards = 0

def pars_excel():
    wb = openpyxl.reader.excel.load_workbook(filename = 'Авито+сбор+разовый+(1).xlsx')
    wb.active = 0
    sheet = wb.active
    urls = []
    def get_urls(maxN, minN):
        urlsF = []
        for i in range(maxN,minN):
            urlsF.append(sheet['B'+str(i)].value)
        return urlsF
    urls += get_urls(6, 21)
    urls += get_urls(22, 37)
    urls += get_urls(38, 52)
    return urls


def get_html(url):
    user = fake_useragent.UserAgent().random
    useragent1 = {'User-Agent': user}

    # proxy_find = choice(proxies)
    # proxy1 = {
    #     'http': 'http://196.216.215.29:56975' + proxy_find,
    #     'https': 'http://196.216.215.29:56975' + proxy_find
    # }
    r = requests.get(url, headers=useragent1) 
    return r


def save_info(el_info):
    wb = openpyxl.load_workbook('ready.xlsx')
    sheet = wb.worksheets[0]
    index = 0 
    for i in el_info:
        index += 1
        sheet.cell(row=el_info['id']+1, column=index).value = el_info[i]
    wb.save('ready.xlsx')


def get_info(url, id):
    global all_cards
    all_cards += 1
    print('card number - ', all_cards)

    html = get_html(base_url + url).text
    soap = BeautifulSoup(html, 'lxml') 

    title = soap.find("div", {"class": "title-info-main"}).text.strip()
    price = soap.find("div", {"class": "item-price-wrapper"}).text.strip()
    el_date = soap.find("div", {"class": "title-info-metadata-item-redesign"}).text.strip()
    kompany = soap.find("div", {"class": 'seller-info-value'}).text.strip()
    user_name = soap.find("div", {"class": 'seller-info-name js-seller-info-name'}).text.strip()
    numberS = soap.find("div", {'class': 'item-view-search-info-redesign'}).text.strip()
    number = ''
    for i in numberS:
        if i == ',':
            break
        else:
            number += i 
    viewed = soap.find("div", {"class": 'title-info-metadata-item title-info-metadata-views'}).text.strip()
    address = soap.find("div", {"class": 'item-address'}).text.strip()
    description = soap.find("div", {"class": 'item-description'}).text.strip()

    el_info = {
        'id': id,
        'title': title,
        'price': price,
        'date': el_date,
        'kompany': kompany,
        'user_name': user_name,
        'number': number,
        'viewed': viewed,
        'address': address,
        'description': description,
        'url': base_url + url
    }
    save_info(el_info)


def get_info_from_page(url):
    html = get_html(url)
    # print(html.text)
    print(html)
    soap = BeautifulSoup(html.text, 'lxml')
    links = soap.find_all('a', class_ = 'link-link-39EVK link-design-default-2sPEv title-root-395AQ iva-item-title-1Rmmj title-listRedesign-3RaU2 title-root_maxHeight-3obWc')

    pog = soap.find_all('span', class_ = 'pagination-item-1WyVp')
    pog = pog[len(pog)-2:len(pog)-1]
    pages = int(pog[0].text)

    for i in range(2, pages):
        print('page ', i)
        html = get_html(url).text
        soap = BeautifulSoup(html, 'lxml')
        linksFPages = soap.find_all('a', class_ = 'link-link-39EVK link-design-default-2sPEv title-root-395AQ iva-item-title-1Rmmj title-listRedesign-3RaU2 title-root_maxHeight-3obWc')
        links.append(linksFPages)

        index = 0
        for link in links:
            index += 1
            get_info(link['href'], index)
        links = []


def finish():
    excel_urls = pars_excel()
    for url in excel_urls:
        get_info_from_page(url) 


finish()





